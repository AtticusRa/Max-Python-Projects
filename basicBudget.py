##Max Friedlander
##10-27-2020
##Basic Budgeting app.
""" Takes in user input and offers basic advice based on what type of budget you would like to have,
if your spending is extreme enough it puts out a warning and new budget. """
"""
TODO:
Create a Spreadsheet/csv that will be uploaded to google sheets. https://realpython.com/openpyxl-excel-spreadsheets-python/#writing-excel-spreadsheets-with-openpyxl
Add option to read and categorize transactions from one csv to another, aka from a bank statement to a budget document https://www.mint.com/mint-categories
Use pre selected designations like housing, rent, food, etc. https://www.quicken.com/blog/budget-categories
Create pretty output of end budget, export to csv file? 
Graphical output to better visualize the budget breakdown https://docs.google.com/spreadsheets/d/1w_tRBJ8ThfSyTctI_Mk2EnKLNbUZBTvjwijDtVLeW10/edit#gid=0
Organize expenses by category so they can fill into csv by category

Add valueError testing to user inputs for spending and income
Compare theoretical ideal spending to current actual spending to improve budget

1. Housing (25-35%)
2. Transportation (10-15)
3. Food (10-15)
4. Utilities (5-10)
5. Insurance (10-25)
6. Medical (5-10)
7. Saving, invest, debt (10-20)
-- non essential
8. Personal (5-10)
9. Rec + entertainment (5-10)
10. misc (5-10)


"""

import re, os
from openpyxl import Workbook, formula


def main():
    name = input('Hi, what is your name? ')
    newBudget = budget(0, 0, name)
    print(f'Hello, {newBudget.name}.')
    newBudget.incomeSources()
    newBudget.expenseSources()
    newBudget.analyzeBudget()
    newBudget.createSheet()


class budget():
    def __init__(self, monthlyIncome, monthlyExpense, name):
        self.monthlyIncome = monthlyIncome
        self.monthlyExpense = monthlyExpense
        self.incomeInput = []
        self.expenseInput = []
        self.incomeCategories = [
            'Paychecks', 'Investments', 'Returned Purchases', 'Bonuses',
            'Interest Income', 'Reimbursements', 'Rental Incomes'
        ]
        self.expenseCategories = [
            'Movies', 'Music', 'Streaming Services/Subscriptions', 'Groceries',
            'Eating Out/Delivery', 'Gifts/Donations', 'Health/Medical', 'Gas',
            'Rent', 'Auto Repair/Transportation', 'Pets', 'Utilities',
            'Debt & Interest Payments', 'Personal Care', 'Clothing',
            'Electronics/Virtual Products', 'Education Expenses', 'Phone',
            'Fees & Charges', 'Travel'
        ]
        self.name = name
        self.wb = Workbook()

    """
    Populate the incomeInput list with user responses, converting all responses to floats. Non used categories use value 0.
     
    
    """

    def incomeSources(self):
        print(
            'Please enter the average amount your household earns from each item every month.\n'
        )
        self.incomeCategories.sort()
        for source in self.incomeCategories:
            self.incomeInput.append(float(input(source + ': ')))
        self.monthlyIncome = sum(self.incomeInput)
        print('Your current monthly reported income is: ' +
              str(self.monthlyIncome))

    #Populate the expenseInput list with user responses, converting all responses to floats. Non used categories use value 0.
    def expenseSources(self):
        print(
            'Please enter the average amount your household spends on each item every month. If you don\'t use this item, enter 0.\n'
        )
        self.expenseCategories.sort()
        for source in self.expenseCategories:
            self.expenseInput.append(float(input(source + ': ')))
        self.monthlyExpense = sum(self.expenseInput)
        print('Your current reported monthly expenses are: ' +
              str(self.monthlyExpense))
        if self.monthlyIncome > self.monthlyExpense:
            print('Great, you are making more than you are spending.')
        else:
            print(
                'You are in the red, but that\'s okay. Let\'s make a great new budget to help out.'
            )

    #Output the budget information to the user and the terminal. Replace with better visual output.
    def displayBudget(self):
        res = '\n'.join(
            '{} {}'.format(x, y)
            for x, y in zip(self.incomeCategories, self.incomeInput))
        res2 = '\n'.join(
            '{} {}'.format(a, b)
            for a, b in zip(self.expenseCategories, self.expenseInput))
        print(f'Hello again, {self.name}, your income sources are: \n\n' + res)
        print('\n')
        print(f'Finally, {self.name}, your expenses are:\n\n' + res2)

    #Creates the workbook sheet that we will use to make the final CSV product.
    def createSheet(self):
        print('Creating Spreadsheet from entered information... ')
        ws = self.wb.active
        ws.title = f"{self.name}\'s Budget"
        for row, entry in enumerate(self.expenseCategories, 1):
            ws.cell(row=row, column=3, value=entry)
        for row, entry in enumerate(self.incomeCategories, 1):
            ws.cell(row=row, column=1, value=entry)
        for row, entry in enumerate(self.incomeInput, 1):
            ws.cell(row=row, column=2, value=entry)
        for row, entry in enumerate(self.expenseInput, 1):
            ws.cell(row=row, column=4, value=entry)

        ws['F6'] = 'Monthly Income:'
        ws['F7'] = 'Monthly Expense:'
        ws['G6'] = self.monthlyIncome
        ws['G7'] = self.monthlyExpense
        ws['G8'] = 'Net:'
        ws['H8'] = self.monthlyIncome - self.monthlyExpense

        self.wb.save(f'{self.name}\'s Budget.xlsx')
        print('Workbook created at ' + str(os.getcwd()))

    #Takes pre set budget % skews and compares them against the actual % skews from the budget.
    def analyzeBudget(self):
        pass


if __name__ == '__main__':
    main()
